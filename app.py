import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
from engine_lib import load_aircraft_dict, save_aircraft_dict, terminate_list

from openpyxl import load_workbook
from engine_lib import addNewEngine, getEngine, getAircraft, editExcel, getCell, addSchedule, getTail, rangeSchedule, addSpare
from engine_lib import PlanShopDate, PlanSchedule, row_for, cleanSchedule, getVisit, updateVisit, find_min_owner, determineOffset

from ExcelRule import RedFillCell, configureFormat
from datetime import datetime, timedelta
import os

from datetime import date
SPARE_FACTOR = 2
CYCLE_PER_DAY = 8

st.set_page_config(page_title="Engine Fleet Staggering", layout="wide")
st.title("Engine Fleet Staggering – Scheduler UI")

uploaded = st.file_uploader("Upload Excel (.xlsx/.xlsm)", type=["xlsx", "xlsm"])

if uploaded: #Uploaded excel file update
    if "excel_bytes" not in st.session_state or st.session_state.get("upload_name") != uploaded.name:
        st.session_state.excel_bytes = uploaded.getvalue()
        st.session_state.upload_name = uploaded.name    

        raw = load_aircraft_dict()
        st.session_state.ListAirCraft = {int(k): v for k, v in raw.items()}

        #st.session_state.ListAirCraft = load_aircraft_dict()   # optional reset on new file
    
    
    
    


    wb = load_workbook(BytesIO(st.session_state.excel_bytes),
                       keep_vba=uploaded.name.endswith(".xlsm"))

    sheet = st.selectbox("Select sheet", wb.sheetnames)
    ws = wb[sheet]
    if "SparePairs" not in st.session_state:
        st.session_state.SparePairs = []
    if "ListAirCraft" not in st.session_state:
        st.session_state.ListAirCraft = {} 
    if "SpareEngineDict" not in st.session_state:
        st.session_state.SpareEngineDict = {}
    listShort = st.session_state.ListAirCraft

    SpareShort = st.session_state.SpareEngineDict

    
    if "Spare1" not in st.session_state:
        st.session_state.Spare1 = []
    

    st.subheader("Current Status")
    
    #add dropped downoption
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        msn = st.number_input("MSN", min_value=0, step=1)
        eng = st.selectbox("Engine Position", ["Eng1", "Eng2"])
    with col2:
        selected_date = st.date_input("Operation start month",value=date.today(),format="DD/MM/YYYY")



    



        #month = st.number_input("Operation start month", min_value=1, max_value=12, step=1, value=1)
    with col3:
        cycle_plan = st.number_input("Target Run", min_value=0, step=1000, value=9000)
        avg_cycle = st.number_input("Average cycles/month", min_value=1, step=1, value=300)

    with col4:
        EngineSerial = st.text_input("Engine S/N", placeholder="e.g. 59D199")
        EngineType = st.selectbox("Engine Option", ["LEAP-1A32", "LEAP-1A26"])
    # Example action buttons
    b1, b2, = st.columns(2)

    month = selected_date.month
    year = selected_date.year
    #listShort[msn]["StartOperation"]  = selected_date
    with b1: #write schedule
        if st.button("Add Aircraft Engine Schedule"):

           
            # e.g. writeSchedule(month, year, cycle_plan, ws, msn, eng, ListAirCraft)
            address = getTail(st.session_state.ListAirCraft)
            TailAdd = msn

            newEntry = addNewEngine(TailAdd, "Eng1", "Eng2") 
            st.session_state.ListAirCraft.update(newEntry)
            save_aircraft_dict(st.session_state.ListAirCraft)
            

            #print(ListAirCraft)
            print(str(address) + "Added New")

            editExcel(address, newEntry, TailAdd, ws, st.session_state.ListAirCraft, EngineSerial)
            #st.write("Updated aircraft dict:", st.session_state.ListAirCraft)

            #Updated file 
            
            
            out = BytesIO()
            wb.save(out)
            
            out.seek(0)
            st.success("MSN " + str(msn) + " Successfully added")
            st.session_state.excel_bytes = out.getvalue()
            #st.write("Updated file size (bytes):", len(st.session_state.updated_excel)) 
           
            #print("Succesful")
            


    with b2:
        if st.button("Clean aircraft list"):
            terminate_list()
            st.session_state.ListAirCraft = {} ##Key 

            st.success("Successfully clean aircraft list")


        out = BytesIO()
        wb.save(out)
        
        out.seek(0)
        #st.success("MSN " + str(msn) + " Successfully added")
        st.session_state.excel_bytes = out.getvalue()

    
    st.write("")
    st.write("")
    
    #MSN control

    st.subheader("Aircraft Information")


    msn_list = list(st.session_state.ListAirCraft.keys())
    if "SpareEngineList" not in st.session_state or not isinstance(st.session_state.SpareEngineList, dict):
        st.session_state.SpareEngineList = {}
    
    



    #SpareEngineUpdate = list(SpareShort)

  
    col5, col6 = st.columns(2)

    First = list(listShort.items())

    Spare1Short = st.session_state.Spare1
    #for msn, rec in First[:2]:       # first 2 aircraft (adjust if needed)
    #    Spare1.append((msn, "Eng1", rec["Eng1"]))
    #    Spare1.append((msn, "Eng2", rec["Eng2"]))

   
    Spare1 = []
    for msn, rec in First[-2:]:
        Spare1.append((msn, "Eng1", rec["Eng1"]))
        Spare1.append((msn, "Eng2", rec["Eng2"]))


    MSN1 = []

    
            
    Spare1 = []

   
    #Schuedule options
    col12, col22, = st.columns(2)
    
   
    SpareEngineUpdate = list(st.session_state.SpareEngineList.keys())

    msn_options = msn_list or ["-- No MSN available --"]
    #st.write("DEBUG rec1 type:", type(SpareEngineUpdate))
    with col5:
        selected_msn = st.selectbox(
        "Select MSN",
        options=msn_options,
        index=0 if msn_options else None
        )
        
        
        selected_spare = st.selectbox("Select Spare", 
        options=SpareEngineUpdate,
        index=0 if msn_list else None)


        if selected_msn != "-- No MSN available --":
            listShort[selected_msn]["StartOperation"] = selected_date
            StartOp = listShort[selected_msn]["StartOperation"]
        else:
            st.warning("No MSN available. Please add/upload aircraft first.")
            st.stop()

        #st.success("Plan applied (replace TODO with your function).")
    
    with col6:
        msnControl1 = st.selectbox("Schedule Option", ["30000", "40000"])
        #if msnControl1 == "30000":
        #   st.success("Plan applied (replace TODO with your function).")
   
    b3, b4  = st.columns(2)
    with col12:
        subEntry = st.selectbox("Schedule Option", ["Single", "Automatic"])
        
        if subEntry == "Automatic":

            optionS, optionE  = st.columns(2)
            with optionS:

                with st.popover("Set automatic start"): #Hover option

                    yearS = st.number_input("Start Year", min_value=2020, max_value=2100, step=1, value=2025)
                    monthS = st.number_input("Start Month", min_value=1, max_value=12, step=1, value=1)

            with optionE:

                with st.popover("Set automatic end"): #Hover option

                    yearE = st.number_input("End Year", min_value=2020, max_value=2100, step=1, value=2025)
                    monthE = st.number_input("End Month", min_value=1, max_value=12, step=1, value=1)

                
    with col22:
        CleanAmount = st.number_input("Clean Schedule", min_value=1, max_value=12, step=1, value=1)


    with b3: 
        if st.button("Add Schedule"):

            if (subEntry == "Single"):

                addSchedule(selected_msn, ws, listShort, cycle_plan, month, year, eng)

                st.success("Single Mode updated")

            if (subEntry == "Automatic"):
                
               
                rangeSchedule(selected_msn, monthS, yearS, monthE, yearE, listShort, ws, cycle_plan, eng) 

                st.success("Automatic Mode updated" + str(SpareEngineUpdate))



            #st.success("Plan applied (replace TODO with your function).")
        
    
        out = BytesIO()
        wb.save(out)
        
        out.seek(0)
        #st.success("MSN " + str(msn) + " Successfully added")
        st.session_state.excel_bytes = out.getvalue()
    
    
    

    
    # Download result
    with b4: 
        if st.button("Clean Schedule"):
            st.success("Plan applied (replace TODO with your function).")
        
    #Update operation date
   


    st.write("")
    st.write("")

    st.subheader("Aircraft Stagging")
    col9, col10, = st.columns(2)
   
    st.subheader("Stagging cycle forecast")
    col31, col32, col33= st.columns(3)
    col41,col42 = st.columns(2)
    #Button stagging
    st1, st2  = st.columns(2)
    st3, st4  = st.columns(2)
    with col9:

        OptionStagging = st.selectbox("Stagging Option", ["Automatic", "Manual"])

        if OptionStagging == "Manual":

            Stagging,  = st.columns(1)

            with Stagging:

                with st.popover("Set Manual start"): #Hover option

                    StaggingYear = st.number_input("Start Year Stagging", min_value=2020, max_value=2100, step=1, value=2025)
                    StaggingMonth = st.number_input("Start Month Stagging", min_value=1, max_value=12, step=1, value=1)


        ShopVisitPurpose = st.selectbox("Shop Visit Scope", ["Engine Performance Restoration 1", "Engine Performance Restoration 2", "LLPs"])

    with col10:
        CleanStagging = st.selectbox("Cleaning Option", ["Manual", "Automatic"])
        
    #Startrt Op
    listShort[selected_msn]["StartOperation"] = selected_date
    StartOp = listShort[selected_msn]["StartOperation"]
    with col31: 
        cycleEGTM = st.number_input("Remaining Cycle EGTM 1", min_value=0, step=1000, value=9000)
        cycleRemainingTarget = st.number_input("Remaining Cycle Target Run 1", min_value=0, step=1000, value=9000)
        cycleBasedOnFan = st.number_input("Lowest LLP remaining Fan 1", min_value=0, step=1000, value=9000)

    with col32: 
        cycleEGTM2 = st.number_input("Remaining Cycle EGTM 2", min_value=0, step=1000, value=9000)
        cycleRemainingTarget2 = st.number_input("Remaining Cycle Target Run 2", min_value=0, step=1000, value=9000)
        cycleBasedOnFan2 = st.number_input("Lowest LLP remaining Fan 2", min_value=0, step=1000, value=9000)

    with col33: 
        cycleEGTMllp = st.number_input("Remaining Cycle EGTM LLP", min_value=0, step=1000, value=9000)
        cycleRemainingTargetllp = st.number_input("Remaining Cycle Target Run LLP", min_value=0, step=1000, value=9000)
        cycleBasedOnFanllp = st.number_input("Lowest LLP remaining Fan LLP", min_value=0, step=1000, value=9000)


    with col41:
        SetFactor = st.number_input("Average cycle per day", min_value=1.0, step=0.1, value=5.0)
        #SetFactor1 = st.number_input("Average cycle per day 2", min_value=1.0, step=0.1, value=6.0)

    #Get forcast date after update
    

    InputShop1 = [cycleEGTM, cycleRemainingTarget, cycleBasedOnFan]
    InputShop2 = [cycleEGTM2, cycleRemainingTarget2, cycleBasedOnFan2]
    InputShopLLp = [cycleEGTMllp, cycleRemainingTargetllp, cycleBasedOnFanllp]

    ListInputForecast = [InputShop1, InputShop2, InputShopLLp]
    #Selected date started 

    #getIndex = getVisit(ShopVisitPurpose) 
    
    
    #for msn, rec in First[:2]:       # first 2 aircraft (adjust if needed)
    #    Spare1.append((msn, "Eng1", rec["Eng1"]))
    #    Spare1.append((msn, "Eng2", rec["Eng2"]))

   
    
    
    with st1: 
        if st.button("Engine Stagging Forecast"):
            st.write("DEBUG appended, now:", st.session_state.Spare1, "rows")
            if OptionStagging == "Automatic":
                PlanSchedule(selected_msn, ws, listShort, 300, eng) #Case nULL no vist
                
                #endDate = selected_date + timedelta(days=Forecast_Delta)

                updateVisit(selected_msn, listShort, ListInputForecast, SetFactor, selected_date, eng)

                
                st.write("MSN " + str(listShort.get(selected_msn)) + " forecast date Successfully added "  )
                st.success(str(StartOp))
                #st.success("Automatic Stagging mode updated " + str(listShort.get(selected_msn)))
                #st.success(str(Spare1))

            if OptionStagging == "Manual":
                PlanShopDate(selected_msn, 6, StaggingMonth, StaggingYear, listShort, ws, eng)
                st.success("Manual Stagging mode updated")
        

        out = BytesIO()
        wb.save(out)
        
        out.seek(0)
        #st.success("MSN " + str(msn) + " Successfully added")
        st.session_state.excel_bytes = out.getvalue()

    #st.write("DEBUG sample aircraft keys:", list(listShort.keys())[:10])
    #st.write("DEBUG selected_msn:", selected_msn, type(selected_msn))
    dictPurpose = {
        "Engine Performance Restoration 1": "FirstVisit", 
        "Engine Performance Restoration 2": "SecondVisit", 
        "LLPs": "ThirdVisit",
    }
    
    with st3:
        #Find initial first earliest forecast:
        #Engine ???? Pair Eng 2: return the earliest (index of)

        if st.button("Finalise schedule"):
            st.write("Spare1 preview:", Spare1)

            #index = find_min_owner(Spare1, "FirstVisit")

            remaining = determineOffset(dictPurpose.get(ShopVisitPurpose), listShort, Spare1)
            st.write("Spare1 preview:",  listShort)

            #st.success(remaining)
            #st.write("Spare1 preview:", listShort)
            #st.success(index)





        out = BytesIO()
        wb.save(out)
        
        out.seek(0)
        #st.success("MSN " + str(msn) + " Successfully added")
        st.session_state.excel_bytes = out.getvalue()



















    st.write("")
    st.write("")
    ## Download output documents 
    download_bytes = st.session_state.get("updated_excel")

    if download_bytes is None:
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        download_bytes = out.getvalue()
    
    st.download_button(
        label="Download updated Excel",
        data=download_bytes,
        file_name="engine_fleet_staggering_updated.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
   
else:

    st.info("Upload an Excel file to begin.")
