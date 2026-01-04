from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font

OOT_YEAR = 2025
ROOT_MONTH = 1
ROOT_COL = 6     # F
ROW = 11

ROOT_ROW = 11          # first MSN block starts here
Current_Row = 51

def RedFillCell(cell, ws):
    
    CellRule = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    #cell_ref = cell if isinstance(cell, str) else cell.coordinate
    
    rule = FormulaRule(
    formula= [str(cell) + ">0"],
    fill=CellRule)
   

    ws.conditional_formatting.add(cell, rule)
    #print(ws.conditional_formatting)
    
def Redfill(cell, ws):

    CellRule = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws[cell].fill = CellRule

def CleanFill(cell, ws):
    ws[cell].fill = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
    ws[cell].font = Font(color="FF000000")  # black text


def configureFormat(cell, currentRow):
    for a in range (ROOT_ROW, currentRow):

        RedFillCell()

def lower(input):
    return 