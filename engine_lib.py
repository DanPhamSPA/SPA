# engine_lib.py
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
import numpy as np
from openpyxl.styles import Alignment,Font
from ExcelRule import RedFillCell, Redfill, CleanFill
import json
import os
from datetime import timedelta

from datetime import date, datetime
AIRCRAFT_JSON = "aircraft.json"
EXPECTED_LOSS = 1800
DESIRED_CYCLE = 9000 
ROOT_YEAR = 2025
ROOT_MONTH = 1
ROOT_COL = 6     # F
ROW = 11
FAN_LLP = 30000
HPC_LLP = 17500
HPT_LLP = 17500
LPT_LLP = 30000 


SHOPVISIT_FACTOR = 6 #6 Months or 180 days

UTILISATION = 3000
SINGLEFL = 1.7 #hours per life cycle 

EXPECTED_CSN  = 9000 #Expected cycle since new 

ROOT_ROW = 11          # first MSN block starts here
BLOCK_HEIGHT = 4       # every MSN takes 3 rows

ENGINE_OFFSET = {
    "Eng1": 0,
    "Eng2": 1,
    "Spacer": 2,   # optional
    "Cycle": 3
}

ShopVisitConvert = {
    "Engine Performance Restoration 1": 0,
    "Engine Performance Restoration 2": 1,
    "LLPs": 2
}

def updateVisit(MSN,listAC, listVisit, SetFactor, selectedDate):


    ShopVisit_days = SHOPVISIT_FACTOR * 30

    #Assumption is that the next start forecas the last of previous shop visit limit + Forecast time 
    
    if MSN not in listAC:
        raise KeyError(f"MSN {MSN} not found in aircraft list")

    Forecast_Delta1 = min(listVisit[0][0], listVisit[0][1], listVisit[0][2])/SetFactor
    Forecast_Delta2 = min(listVisit[1][0], listVisit[1][1], listVisit[1][2])/SetFactor
    Forecast_Delta3 = min(listVisit[2][0], listVisit[2][1], listVisit[2][2])/SetFactor


    listAC[MSN]["FirstVisit"] = selectedDate + timedelta(days=Forecast_Delta1+ ShopVisit_days)
    listAC[MSN]["SecondVisit"] =  listAC[MSN]["FirstVisit"] + timedelta(days=Forecast_Delta2 + ShopVisit_days)
    listAC[MSN]["ThirdVisit"] = listAC[MSN]["SecondVisit"] + timedelta(days=Forecast_Delta3+ ShopVisit_days)
    
    



    return
def _json_default(o):
    if isinstance(o, (date, datetime)):
        return o.isoformat()   # e.g. "2026-06-13"
    raise TypeError(f"Object of type {type(o).__name__} is not JSON serializable")



def getVisit(entry):
    return ShopVisitConvert.get(entry)

def terminate_list(path=AIRCRAFT_JSON):
    open(path, 'w').close()

def save_aircraft_dict(data, path=AIRCRAFT_JSON):
     with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=_json_default)


def load_aircraft_dict(path=AIRCRAFT_JSON):
    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            json.dump({}, f)
        return {}

    # Load existing file
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except json.JSONDecodeError:
        # Recover from corrupted / empty file
        with open(path, "w", encoding="utf-8") as f:
            json.dump({}, f)
        return {}

def msn_index(msn: int, aircraft_dict: dict) -> int:
    """Return row-block index based on insertion order of dict keys."""
    return list(aircraft_dict.keys()).index(msn)

def row_for(msn: int, engine: int, aircraft_dict: dict) -> int:
    i = msn_index(msn, aircraft_dict)

    return ROOT_ROW + BLOCK_HEIGHT * i + ENGINE_OFFSET[engine]

def getTail(list):
    length = len(list)
    return int(length*4 + ROOT_ROW)
    
def addNewEngine(MSN, Eng1, Eng2):

    #Schedule1 = np.zeros((10, 12))
    #Schedule2 = np.zeros((10, 12))

    newEntry = {MSN:{Eng1:{"CycleR":100, "Schedule": 1},"StartOperation":0, 
                     Eng2:{"CycleR":0, "Schedule": 1}, "ShopVisit": 1, "FirstVisit":0,
                     "SecondVisit":0, "ThirdVisit":0}}

    #Schedule1[0][0] = 120
    
    return newEntry

def getEngine(MSN, Postion, ListAirCraft):
    if Postion == "Eng1":
        return ListAirCraft.get(MSN).get("Eng1")
    else: 
        return ListAirCraft.get(MSN).get("Eng2")
    
def getAircraft(MSN, ListAirCraft):
    
    return ListAirCraft.get(MSN)

def editExcel(address, newEntry, tail, ws, list, EngineSerial):
    #print(tail)
    #result = list.get(tail).get("Eng1").get("Schedule")[0][0]
    #print(result)
    ws["D" + str(address)] = tail
    ws["D" + str(address)].alignment = Alignment(horizontal="center", vertical="center")
    ws["D" + str(address)].font = Font(bold=True)

    #Engine layout
    ws["C" + str(address)] = 1
    ws["C" + str(address)].alignment = Alignment(horizontal="center", vertical="center")
    ws["C" + str(address)].font = Font(bold=True)

    ws["C" + str(address + 1)] = 2 
    ws["C" + str(address+1)].alignment = Alignment(horizontal="center", vertical="center")
    ws["C" + str(address+1)].font = Font(bold=True)

    ws["D" + str(address+ 2) ] = "Remaining cycles"
    ws["D" + str(address+2 )].alignment = Alignment(horizontal="center", vertical="center")
    ws["D" + str(address+2)].font = Font(bold=True)
    
    ws["D" + str(address + 3)] = "Remaining cycles"
    ws["D" + str(address + 3)].alignment = Alignment(horizontal="center", vertical="center")
    ws["D" + str(address + 3)].font = Font(bold=True)


    ##Total Cycle 
    ws["B" + str(address+ 2) ] = "Total Cycle 1"
    ws["B" + str(address+2 )].alignment = Alignment(horizontal="center", vertical="center")
    ws["B" + str(address+2)].font = Font(bold=True)
    
    ws["B" + str(address + 3)] = "Total Cycle 2"
    ws["B" + str(address + 3)].alignment = Alignment(horizontal="center", vertical="center")
    ws["B" + str(address + 3)].font = Font(bold=True)
    

    subEntry = "none"
    while subEntry:

        if (subEntry.lower() == "none"):
            break

        if (subEntry.lower() == "addschedule"):
            Month = input("Month: ")
            Year = input("Year: ")
            CyclePlanned = input("CyclePlan: ")

            EngineNumber = input("EngineNumber: ")

            if EngineNumber == "1":
                writeSchedule(Month, Year, CyclePlanned, ws, tail, "Eng1", list)
            else: 
                writeSchedule(Month, Year, CyclePlanned, ws, tail, "Eng2", list)


        subEntry = input("Edit Option: ")
    
def writeSchedule(Month, Year, cycleRan, ws, MSN, Engine, list):
    rootAddress = "F11" #Row 11 column F 

    #Start from F9 Assume Start Year 2025 
    #Jan 2025 = F11 
    #Jan 2026 = (F11 + 12 * (year-rootyear) + Month)

    writeOffset = month_offset(int(Year), int(Month))  # Jan 2026
    # → 12

    #Write the the schedule 
    col = ROOT_COL + writeOffset
    rowPosition = row_for(MSN, Engine, list)

    
    ws.cell(row=rowPosition, column=col).value = cycleRan


    return

def month_offset(year, month, root_year=2025, root_month=1):

    return (year - root_year) * 12 + (month - root_month)

def getCell(Month, Year, ws, MSN, Engine, list):

    writeOffset = month_offset(int(Year), int(Month))  # Jan 2026

    col = ROOT_COL + writeOffset
    rowPosition = row_for(MSN, Engine, list)

    return col, rowPosition

def addSchedule(tail, ws, list, CyclePlanned, Month, Year, EngineNumber):


    #Month = input("Month: ")
    #Year = input("Year: ")
    #CyclePlanned = input("CyclePlan: ")

    #EngineNumber = input("EngineNumber: ")

    if EngineNumber == "Eng1":
        writeSchedule(Month, Year, CyclePlanned, ws, tail, "Eng1", list)
    else: 
        writeSchedule(Month, Year, CyclePlanned, ws, tail, "Eng2", list)

def rangeSchedule(getMsn, sM, sY, endM, endY, list, ws, cycleRan, eng):
    #start row/column 
    print("Operated")
    initialOff = month_offset(int(sY), int(sM))
    initalCol = ROOT_COL + initialOff

    endOff = month_offset(int(endY), int(endM))
    endCol = ROOT_COL + endOff

    #engOp = input("Engine Option ")

    rowPosition = row_for(getMsn, eng, list)
    
    #get column

    for i in range (initalCol, endCol):
        ws.cell(row=rowPosition, column=i).value = cycleRan

    print("Finished")
def PlanShopDate(MSN, Duration, sM, sY, list, ws, eng):
    #print(Duration)
    initialOff = month_offset(int(sY), int(sM))
    #print(initialOff)

    initalCol = ROOT_COL + initialOff

    #print("Initial" + str(initalCol))
    #endOff = month_offset(int(sY), int())

    endCol = initalCol + Duration 

    #print("Initial" + str(endCol))
    #engOp = input("Engine Option ")

    rowPosition = row_for(MSN, eng, list)
    #print(rowPosition)

    for i in range (initalCol, endCol):
        col_letter = get_column_letter(i)
        cell = col_letter + str(rowPosition)
        
        Redfill(cell, ws)

    return

def PlanSchedule(MSN, ws, list, averageCycle, Eng):
    firstStaggering = DESIRED_CYCLE - EXPECTED_LOSS

    MonthToStagger = int(firstStaggering/averageCycle)
    #print(MonthToStagger)
    
    Start = ROOT_COL + MonthToStagger
    #print(Start)
    rowPosition = row_for(MSN, Eng, list)
    #print(rowPosition)

    for i in range (Start, Start+6):

        col_letter = get_column_letter(i)

        cell = f"{col_letter}{rowPosition}"
        #print(cell)

        Redfill(cell, ws)

def cleanSchedule(MSN, Duration, sM, sY, list, ws):
    #print(Duration)
    initialOff = month_offset(int(sY), int(sM))
    #print(initialOff)

    initalCol = ROOT_COL + initialOff

    #print("Initial" + str(initalCol))
    #endOff = month_offset(int(sY), int())

    endCol = initalCol + Duration 

    #print("Initial" + str(endCol))
    engOp = input("Engine Option ")

    rowPosition = row_for(MSN, engOp, list)
    #print(rowPosition)

    for i in range (initalCol, endCol):
        col_letter = get_column_letter(i)
        cell = col_letter + str(rowPosition)
        
        CleanFill(cell, ws)

    return


#End